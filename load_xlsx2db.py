import sys
import os
import argparse
from openpyxl import load_workbook
# import pymssql
import xml.etree.ElementTree as ET
import re
import logging

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


def load_config(config_file):
    tree = ET.parse(config_file)
    root = tree.getroot()

    db_prop = {}
    for child in root.iter('target-db'):
        # only one target
        db_prop['name'] = child.attrib['name']
        db_prop['type'] = child.attrib['type']
        for p in child:
            db_prop[p.tag] = p.text

    xls_prop = {}
    for child in root.iter('source-folder'):
        # only one source
        xls_prop['name'] = child.attrib['name']
        xls_prop['type'] = child.attrib['type']
        for p in child:
            xls_prop[p.tag] = p.text

    xref_prop = {}
    for child in root.iter('load-map'):
        # only one load-map
        for xc in child:
            if xc.tag == 'source':
                src = {}
                for p in xc:
                    if p.tag == 'file':
                        src['file-name'] = p.attrib['name']
                    elif p.tag == 'tab-names':
                        src['tab-name'] = p.attrib['name']
                    elif p.tag == 'header-range':
                        src['header'] = {'from': p.attrib['from'], 'to': p.attrib['to']}
                    elif p.tag == 'data-range':
                        src['data'] = {'from': p.attrib['from'], 'to': p.attrib['to']}
                    elif p.tag == 'anchors':
                        src['anchors'] = []
                        for a in p.iter('anchor'):
                            src['anchors'].append({'name': a.attrib['name'], 'at': a.attrib['at']})
                xref_prop["source"] = src
            elif xc.tag == 'target':
                tgt = {}
                for p in xc:
                    if p.tag == 'table':
                        tgt['table-name'] = p.attrib['name']
                xref_prop["target"] = tgt
            elif xc.tag == 'columns':
                cols = {}
                for p in xc.iter('column'):
                    cols[p.attrib['source'].strip()] = {
                        'name': p.attrib['target'],
                        'type': p.attrib['type'],
                        'role': p.attrib['role'] if 'role' in p.attrib else None
                    }
                xref_prop['columns'] = cols

    return {
        "database": db_prop,
        "folder": xls_prop,
        "xref": xref_prop
    }


def insert_data(cfg, headers, data, anchors):
    # insert into table (col, col) values (val, val)
    for row in data:
        col_names = []
        col_values = []
        for c in range(0, len(row)):
            hdr = headers[c]
            if hdr in cfg['xref']['columns'].keys():
                col_names.append(cfg['xref']['columns'][hdr]['name'])
                if cfg['xref']['columns'][hdr]['type'] == 'text':
                    col_values.append("'" + str(row[c]).strip() + "'")
                elif cfg['xref']['columns'][hdr]['type'] == 'number':
                    # TODO: strip non-numeric chars
                    col_values.append(row[c])
                elif cfg['xref']['columns'][hdr]['type'] == 'datetime':
                    # TODO: convert to datetime
                    col_values.append(str(row[c]))
        for ahr in anchors:
            col_names.append(cfg['xref']['columns'][ahr['name']]['name'])
            col_values.append("'" + ahr['value'] + "'")
        insert_stmt = [
            "insert into ",
            cfg['xref']['target']['table-name'],
            " (",
            ",".join([n for n in col_names]),
            ") values (",
            ",".join([str(v) for v in col_values]),
            ");"
        ]
        print(''.join(insert_stmt))


def load_data(cfg):

    for root, dirs, files in os.walk(cfg['folder']['path']):
        for fname in files:
            if re.search(cfg['xref']['source']['file-name'], fname) is not None:

                xls_file_path = os.path.join(root, fname)
                xf = load_workbook(xls_file_path, data_only=True)

                headers = []
                data = []
                anchors = []

                for sn in xf.sheetnames:
                    if re.search(cfg['xref']['source']['tab-name'], sn) is not None:
                        sheet = xf[sn]
                        # anchors
                        if len(anchors) == 0:
                            for a in cfg['xref']['source']['anchors']:
                                for row in sheet[a['at']:a['at']]:
                                    for cell in row:
                                        anchors.append({'name': a['name'], 'value': cell.value.strip()})

                        # headers
                        if len(headers) == 0:
                            for row in sheet[cfg['xref']['source']['header']['from']:cfg['xref']['source']['header']['to']]:
                                for cell in row:
                                    headers.append(cell.value.strip())

                        # values
                        cell_min_index = cfg['xref']['source']['data']['from']
                        row_min_index = int(cell_min_index[1:])
                        cell_max_index = cfg['xref']['source']['data']['to']
                        row_max_index = row_min_index
                        if re.search('[a-zA-Z]{1}[0-9]+\\+', cell_max_index) is not None:
                            cell_max_index = cell_max_index[:len(cell_max_index)-1] + "00"
                            row_max_index = int(cell_max_index[1:])
                        row_max_count = row_max_index - row_min_index + 1
                        row_count = 0
                        for row in sheet[cell_min_index:cell_max_index]:
                            rcd = []
                            is_row_empty = True
                            for cell in row:
                                is_row_empty = cell.value is None
                                rcd.append(cell.value)
                            if not is_row_empty:
                                data.append(rcd)
                            else:
                                # end the loading when all cells are empty
                                break
                            row_count = row_count + 1
                            # end the loading when the max count is reached
                            if row_count > row_max_count:
                                break

                        # insert data
                        insert_data(cfg, headers, data, anchors)

                xf = None


def load_data_set(config_file):

    cfg = load_config(config_file)

    load_data(cfg)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Load data from xlsx to database')
    parser.add_argument('-c', '--cfg', help='Config File', required=False,
                        default=r'C:\Users\Chen.Liang\Projects\notebook\HenryRes\config-Completion_Summary.xml')
                        # default=r'C:\Users\Chen.Liang\Projects\notebook\HenryRes\config-deviation_survey.xml')

    params = parser.parse_args()

    load_data_set(params.cfg)
